"""Repositorio encargado de la interacción con los archivos Excel."""

from __future__ import annotations

import io
from pathlib import Path
from typing import Dict, List

import openpyxl
import pandas as pd

from asesorias_app import config


def normalize_registro_df(df: pd.DataFrame) -> pd.DataFrame:
    """Garantiza columnas esperadas y descarta campos obsoletos."""
    df = df.copy()
    aliases = getattr(config, "COLUMN_ALIASES", {})
    if aliases:
        rename_map = {}
        for alias, canonical in aliases.items():
            if alias not in df.columns:
                continue
            if canonical in df.columns:
                df = df.drop(columns=[alias], errors="ignore")
            else:
                rename_map[alias] = canonical
        if rename_map:
            df = df.rename(columns=rename_map)
    removed = getattr(config, "REMOVED_REGISTRO_COLUMNS", [])
    if removed:
        df = df.drop(columns=removed, errors="ignore")

    for col in config.REGISTRO_COLUMNS:
        if col not in df.columns:
            df[col] = None

    ordered_cols = config.REGISTRO_COLUMNS + [c for c in df.columns if c not in config.REGISTRO_COLUMNS]
    return df[ordered_cols]


class ExcelRepository:
    """Encapsula el manejo de lectura y escritura en Excel."""

    def __init__(self, db_path: Path | None = None, template_path: Path | None = None) -> None:
        self.db_path = Path(db_path or config.DB_PATH)
        self.template_path = Path(template_path or config.TEMPLATE_PATH)
        config.DATA_DIR.mkdir(parents=True, exist_ok=True)

    # Listas y dataframes -------------------------------------------------
    def load_lists(self) -> dict:
        wb = openpyxl.load_workbook(self.template_path, data_only=True, keep_vba=True)
        ws = wb[config.SHEET_LISTAS]

        def extract_col(col_idx: int, start_row: int = 2):
            vals = []
            for row_idx in range(start_row, ws.max_row + 1):
                value = ws.cell(row_idx, col_idx).value
                if value is None:
                    continue
                text = str(value).strip()
                if text:
                    vals.append(text)
            unique: List[str] = []
            seen = set()
            for item in vals:
                if item not in seen:
                    unique.append(item)
                    seen.add(item)
            return unique

        headers: Dict[str, int] = {}
        for col in range(1, ws.max_column + 1):
            header = ws.cell(1, col).value
            if header:
                headers[str(header).strip()] = col

        lists = {name: extract_col(idx) for name, idx in headers.items()}
        df_fac = pd.read_excel(self.template_path, sheet_name=config.SHEET_FACULTADES)
        df_prog = pd.read_excel(self.template_path, sheet_name=config.SHEET_PROGRAMAS)

        return {"lists": lists, "df_fac": df_fac, "df_prog": df_prog}

    # Registro principal ---------------------------------------------------
    def ensure_db(self) -> None:
        if not self.db_path.exists():
            df_template = pd.read_excel(self.template_path, sheet_name=config.SHEET_REGISTRO)
            df_template = normalize_registro_df(df_template)
            df_empty = df_template.iloc[0:0].copy()
            df_empty.to_excel(self.db_path, index=False)

    def load_registro(self) -> pd.DataFrame:
        self.ensure_db()
        df = pd.read_excel(self.db_path)
        df = normalize_registro_df(df)
        if "Fecha" in df.columns:
            df["Fecha"] = pd.to_datetime(df["Fecha"], errors="coerce")
        return df

    def save_registro(self, df: pd.DataFrame) -> None:
        df = normalize_registro_df(df)
        tmp = self.db_path.with_suffix(".tmp.xlsx")
        df.to_excel(tmp, index=False)
        tmp.replace(self.db_path)

    def download_current_excel_bytes(self) -> bytes:
        df = self.load_registro()
        bio = io.BytesIO()
        with pd.ExcelWriter(bio, engine="xlsxwriter") as writer:
            df.to_excel(writer, index=False, sheet_name="Registro")
        bio.seek(0)
        return bio.read()

    def build_bulk_template(self, columns: List[str]) -> bytes:
        df = pd.DataFrame(columns=columns)
        bio = io.BytesIO()
        with pd.ExcelWriter(bio, engine="xlsxwriter") as writer:
            df.to_excel(writer, index=False, sheet_name="Plantilla")
        bio.seek(0)
        return bio.read()
