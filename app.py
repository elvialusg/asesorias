"""Tablero de asesorías (tesis) con Streamlit."""

from __future__ import annotations

import io
import os
from datetime import date

import openpyxl
import pandas as pd
import streamlit as st

from asesorias_core import init_db
from asesorias_core.config import DATA_DIR, settings
from asesorias_core.constants import (
    COL_ASESOR_METODOLOGICO_DETALLE,
    COL_ASESOR_RECURSOS,
    COL_APROBACION_SIMILITUD,
    COL_CEDULA,
    COL_ESCANEADO_TURNITIN,
    COL_FECHA,
    COL_HISTORIAL_ASESORIAS,
    COL_HISTORIAL_ASESOR_METODOLOGICO,
    COL_HISTORIAL_ASESOR_RECURSOS,
    COL_HISTORIAL_FECHAS,
    COL_MODALIDAD_PROGRAMA,
    COL_NOMBRE_ASESORIA,
    COL_NOMBRE_FACULTAD,
    COL_NOMBRE_PROGRAMA,
    COL_NOMBRE_USUARIO,
    COL_OBSERVACIONES,
    COL_OK_REFERENCISTAS,
    COL_OK_SERVICIOS,
    COL_PAZ_Y_SALVO,
    COL_PORCENTAJE_SIMILITUD,
    COL_REVISION_INICIAL,
    COL_REVISION_PLANTILLA,
    COL_TITULO_TRABAJO,
    COL_TOTAL_ASESORIAS,
)
from asesorias_core.services import ExcelService, RegistroService
from asesorias_core.utils import norm_str, normalize_fac_name

st.set_page_config(page_title="Tablero Asesorías (Tesis)", layout="wide")

init_db()
registro_service = RegistroService()
excel_service = ExcelService()

TEMPLATE_PATH = settings.template_path
SHEET_FAC = "Data_Facultades"
SHEET_PROG = "Data_Programas"
SHEET_LISTAS = "Insumo_Listas_Desplegables"

os.makedirs(os.fspath(DATA_DIR), exist_ok=True)

st.markdown(
    """
    <style>
      :root{--brand:#1B8F4B;--brand-2:#146B39;--bg:#F6F7F9;--card:#FFFFFF;--muted:#6B7280;--border:#E5E7EB;}
      .stApp { background: var(--bg); }
      .muted { color: var(--muted); }
      div.stButton > button[kind="primary"]{background: var(--brand) !important;border: 1px solid var(--brand-2) !important;color: white !important;border-radius: 10px !important;padding: 0.55rem 0.9rem !important;}
      div.stButton > button[kind="secondary"]{border-radius: 10px !important;}
      .card{background: var(--card);border: 1px solid var(--border);border-radius: 14px;padding: 16px 16px 8px 16px;box-shadow: 0 1px 2px rgba(0,0,0,0.05);margin-bottom: 12px;}
      button[data-baseweb="tab"][aria-selected="true"]{ color: var(--brand) !important; }
      [data-testid="stDataFrame"]{border-radius: 14px;overflow: hidden;border: 1px solid var(--border);background: white;}
    </style>
    """,
    unsafe_allow_html=True,
)


@st.cache_data(show_spinner=False)
def load_lists(template_path: str) -> dict:
    wb = openpyxl.load_workbook(template_path, data_only=True, keep_vba=True)
    ws = wb[SHEET_LISTAS]

    def extract_col(col_idx: int, start_row: int = 2):
        vals = []
        for r in range(start_row, ws.max_row + 1):
            value = ws.cell(r, col_idx).value
            if value is None:
                continue
            text = str(value).strip()
            if text:
                vals.append(text)
        out, seen = [], set()
        for val in vals:
            if val not in seen:
                out.append(val)
                seen.add(val)
        return out

    headers = {}
    for c in range(1, ws.max_column + 1):
        header = ws.cell(1, c).value
        if header:
            headers[str(header).strip()] = c

    lists = {name: extract_col(col_idx) for name, col_idx in headers.items()}
    df_fac = pd.read_excel(template_path, sheet_name=SHEET_FAC)
    df_prog = pd.read_excel(template_path, sheet_name=SHEET_PROG)

    fac_names = df_fac["Nombre_Facultad"].dropna().astype(str).tolist()
    fac_norm_map = {normalize_fac_name(n): n for n in fac_names}

    return {"lists": lists, "df_fac": df_fac, "df_prog": df_prog, "fac_norm_map": fac_norm_map}


def load_registro() -> pd.DataFrame:
    return registro_service.get_dataframe()


def download_current_excel_bytes() -> bytes:
    df = load_registro().copy()
    return excel_service.dataframe_to_bytes(df)


def bulk_template_bytes(columns: list[str]) -> bytes:
    return excel_service.build_template(columns)


def bulk_import(df_upload: pd.DataFrame):
    df_norm = excel_service.normalize_upload_df(df_upload)
    registro_service.bulk_import(df_norm)


def history_table_from_row(row: pd.Series) -> pd.DataFrame:
    return registro_service.history_table_from_row(row)


def delete_history_item(record_id: int, position: int) -> bool:
    return registro_service.delete_historial_item(record_id, position)


def all_widget_keys():
    base = [
        "facultad",
        "programa",
        "cedula",
        "nombre_usuario",
        "titulo",
        "fecha",
        "rev_inicial",
        "rev_plantilla",
        "ok_ref",
        "ok_serv",
        "esc_turnitin",
        "similitud",
        "aprob_sim",
        "obs",
        "paz_y_salvo",
    ]
    n = int(st.session_state.get("asesorias_n", 1))
    for i in range(n):
        base += [
            f"asesor_rec_{i}",
            f"nombre_asesoria_{i}",
            f"modalidad_{i}",
            f"asesor_metodologico_{i}",
            f"modalidad2_{i}",
        ]
    return base


def schedule_reset():
    st.session_state["reset_pending"] = True
    st.rerun()


def program_faculty_column(df_prog: pd.DataFrame) -> str:
    candidates = ["Codigo_Facultad", "Código_Facultad", "Códígo_Facultad", "C�digo_Facultad"]
    for col in candidates:
        if col in df_prog.columns:
            return col
    return df_prog.columns[0]


def perform_reset(meta: dict):
    for key in all_widget_keys():
        st.session_state.pop(key, None)

    st.session_state["asesorias_n"] = 1
    st.session_state["cedula"] = ""
    st.session_state["nombre_usuario"] = ""
    st.session_state["titulo"] = ""
    st.session_state["fecha"] = date.today()
    st.session_state["ok_ref"] = ""
    st.session_state["ok_serv"] = ""
    st.session_state["obs"] = ""
    st.session_state["similitud"] = 0.0
    st.session_state["paz_y_salvo"] = "EN PROCESO"

    fac_list = meta["df_fac"]["Nombre_Facultad"].dropna().astype(str).tolist()
    st.session_state["facultad"] = fac_list[0] if fac_list else ""
    if fac_list:
        df_fac = meta["df_fac"]
        df_prog = meta["df_prog"]
        fac_code = int(df_fac.loc[df_fac["Nombre_Facultad"] == fac_list[0], "Codigo_Facultad"].iloc[0])
        prog_field = program_faculty_column(df_prog)
        progs = df_prog.loc[df_prog[prog_field] == fac_code, "Nombre_Programa"].dropna().astype(str).tolist()
        st.session_state["programa"] = progs[0] if progs else ""
    else:
        st.session_state["programa"] = ""

    lists = meta["lists"]
    st.session_state["rev_inicial"] = (lists.get("Revisión Inicial") or [""])[0]
    rev_pl = lists.get("Revisión de Plantilla") or lists.get("Revisión plantilla") or [""]
    st.session_state["rev_plantilla"] = rev_pl[0] if rev_pl else ""
    st.session_state["esc_turnitin"] = (lists.get("Escaneado Turnitin") or [""])[0]
    st.session_state["aprob_sim"] = (lists.get("Aprobados PyS") or [""])[0]

    st.session_state["asesor_rec_0"] = (lists.get("Asesor_Recursos_Académicos") or [""])[0]
    st.session_state["nombre_asesoria_0"] = (lists.get("Nombre_Asesoría") or [""])[0]
    st.session_state["modalidad_0"] = (lists.get("Modalidad_Asesoría") or ["Virtual"])[0]
    st.session_state["asesor_metodologico_0"] = ""
    st.session_state["modalidad2_0"] = ""

    st.session_state["reset_pending"] = False


def ensure_dynamic_defaults(meta: dict):
    lists = meta["lists"]
    n = int(st.session_state.get("asesorias_n", 1))
    for i in range(n):
        st.session_state.setdefault(f"asesor_rec_{i}", (lists.get("Asesor_Recursos_Académicos") or [""])[0])
        st.session_state.setdefault(f"nombre_asesoria_{i}", (lists.get("Nombre_Asesoría") or [""])[0])
        st.session_state.setdefault(f"modalidad_{i}", (lists.get("Modalidad_Asesoría") or ["Virtual"])[0])
        st.session_state.setdefault(f"asesor_metodologico_{i}", "")
        st.session_state.setdefault(f"modalidad2_{i}", "")


def add_asesoria():
    st.session_state["asesorias_n"] = int(st.session_state.get("asesorias_n", 1)) + 1


def remove_asesoria():
    n = int(st.session_state.get("asesorias_n", 1))
    if n <= 1:
        return
    last = n - 1
    for key in [
        f"asesor_rec_{last}",
        f"nombre_asesoria_{last}",
        f"modalidad_{last}",
        f"asesor_metodologico_{last}",
        f"modalidad2_{last}",
    ]:
        st.session_state.pop(key, None)
    st.session_state["asesorias_n"] = n - 1


def autofill_by_cedula(meta: dict):
    ced = norm_str(st.session_state.get("cedula"))
    if not ced:
        return
    match = registro_service.find_student(cedula=ced, nombre=None)
    if not match:
        return

    st.session_state["nombre_usuario"] = str(match.get(COL_NOMBRE_USUARIO) or "")
    st.session_state["titulo"] = str(match.get(COL_TITULO_TRABAJO) or "")
    st.session_state["paz_y_salvo"] = str(match.get(COL_PAZ_Y_SALVO) or "EN PROCESO")

    fac_norm = str(match.get(COL_NOMBRE_FACULTAD) or "").strip()
    fac_display = meta["fac_norm_map"].get(fac_norm)
    if fac_display:
        st.session_state["facultad"] = fac_display
        prog = str(match.get(COL_NOMBRE_PROGRAMA) or "").strip()
        if prog:
            st.session_state["programa"] = prog


# -------------------------------------------------------------------------------
# App
# -------------------------------------------------------------------------------

if not os.path.exists(TEMPLATE_PATH):
    st.error(f"No encuentro el template en: {TEMPLATE_PATH}\n\n✔️ Debe existir: data/Control_Asesorias_Tesis_template.xlsm")
    st.stop()

meta = load_lists(TEMPLATE_PATH)
lists = meta["lists"]
df_fac = meta["df_fac"]
df_prog = meta["df_prog"]

st.title("Asesorías • Trabajo de grado")
tab1, tab2, tab3 = st.tabs(["✚ Registrar asesoría", "🔎 Consultar usuario", "📂 Carga masiva"])

with tab1:
    st.markdown('<div class="card">', unsafe_allow_html=True)
    if st.session_state.get("reset_pending", False) or "init_done" not in st.session_state:
        perform_reset(meta)
        st.session_state["init_done"] = True
    ensure_dynamic_defaults(meta)

    colA, colB = st.columns([1.15, 0.85], gap="large")
    with colA:
        st.subheader("Formulario de registro")
        fac_names = df_fac["Nombre_Facultad"].dropna().astype(str).tolist()
        fac_display = st.selectbox("Facultad", fac_names, index=0, key="facultad")
        fac_code = int(df_fac.loc[df_fac["Nombre_Facultad"] == fac_display, "Codigo_Facultad"].iloc[0])
        prog_field = program_faculty_column(df_prog)
        progs = df_prog.loc[df_prog[prog_field] == fac_code, "Nombre_Programa"].dropna().astype(str).tolist() or [""]
        prog_display = st.selectbox("Programa", progs, index=0, key="programa")
        c1, c2 = st.columns(2)
        with c1:
            cedula = st.text_input("Cédula", placeholder="Ej: 1032331000", key="cedula", on_change=lambda: autofill_by_cedula(meta))
        with c2:
            nombre_usuario = st.text_input("Nombre del usuario", placeholder="Ej: Juan Pérez", key="nombre_usuario")
        titulo = st.text_input("Título trabajo de grado", key="titulo")
        fecha = st.date_input("Fecha", key="fecha")

        st.markdown("### Asesorías a registrar")
        n = int(st.session_state.get("asesorias_n", 1))
        cadd, crem, _ = st.columns([0.25, 0.25, 0.5])
        with cadd:
            st.button("✚ Agregar asesoría", on_click=add_asesoria, key="btn_add_asesoria")
        with crem:
            st.button("➖ Quitar última", on_click=remove_asesoria, disabled=(n <= 1), key="btn_remove_asesoria")

        asesorias_payload = []
        for i in range(n):
            with st.expander(f"Asesoría #{i + 1}", expanded=(i == 0)):
                asesor_rec_i = st.selectbox("Asesor Recursos Académicos", lists.get("Asesor_Recursos_Académicos", [""]), key=f"asesor_rec_{i}")
                nombre_asesoria_i = st.selectbox("Nombre de la asesoría", lists.get("Nombre_Asesoría", [""]), key=f"nombre_asesoria_{i}")
                modalidad_i = st.selectbox("Modalidad", lists.get("Modalidad_Asesoría", ["Virtual", "Presencial"]), key=f"modalidad_{i}")
                st.markdown("**Asesor metodológico**")
                asesor_met_i = st.text_input("Asesor metodológico (digita el nombre)", placeholder="Ej: Carlos Rodríguez", key=f"asesor_metodologico_{i}")
                modalidad2_i = st.selectbox("Modalidad asesoría ", ["", "Virtual", "Presencial"], key=f"modalidad2_{i}")

                if asesor_met_i and modalidad2_i:
                    campo_h_i = f"{asesor_met_i.strip()}, {modalidad2_i}"
                elif asesor_met_i:
                    campo_h_i = asesor_met_i.strip()
                else:
                    campo_h_i = None

                asesorias_payload.append(
                    {
                        COL_ASESOR_RECURSOS: norm_str(asesor_rec_i),
                        COL_NOMBRE_ASESORIA: norm_str(nombre_asesoria_i),
                        COL_MODALIDAD_PROGRAMA: norm_str(modalidad_i),
                        COL_ASESOR_METODOLOGICO_DETALLE: norm_str(campo_h_i),
                        COL_FECHA: pd.to_datetime(fecha),
                    }
                )

        c3, c4 = st.columns(2)
        with c3:
            rev_inicial = st.selectbox("Revisión inicial", lists.get("Revisión Inicial", [""]), key="rev_inicial")
            rev_pl_opts = lists.get("Revisión de Plantilla") or lists.get("Revisión plantilla") or [""]
            rev_plantilla = st.selectbox("Revisión plantilla", rev_pl_opts, key="rev_plantilla")
            ok_ref = st.text_input("Ok_Referencistas", placeholder="Ej: listo / pendiente", key="ok_ref")
        with c4:
            ok_serv = st.text_input("OK_Servicios", placeholder="Ej: listo / pendiente", key="ok_serv")
            esc_turnitin = st.selectbox("Escaneado Turnitin", lists.get("Escaneado Turnitin", [""]), key="esc_turnitin")
            similitud = st.number_input("% similitud", min_value=0.0, max_value=100.0, step=0.1, key="similitud")

        aprob_sim = st.selectbox("Aprobación similitud", lists.get("Aprobados PyS", [""]), key="aprob_sim")
        obs = st.text_area("Observaciones", height=120, key="obs")
        paz_y_salvo = st.selectbox("Estudiante apto para paz y salvo", ["EN PROCESO", "SI", "NO"], index=0, key="paz_y_salvo")

        principal = asesorias_payload[0] if asesorias_payload else {}
        fac_norm = normalize_fac_name(fac_display)
        base_row = {
            COL_NOMBRE_FACULTAD: fac_norm,
            COL_NOMBRE_PROGRAMA: prog_display,
            COL_CEDULA: norm_str(cedula),
            COL_NOMBRE_USUARIO: norm_str(nombre_usuario),
            COL_ASESOR_RECURSOS: principal.get(COL_ASESOR_RECURSOS),
            COL_NOMBRE_ASESORIA: principal.get(COL_NOMBRE_ASESORIA),
            COL_MODALIDAD_PROGRAMA: principal.get(COL_MODALIDAD_PROGRAMA),
            COL_ASESOR_METODOLOGICO_DETALLE: principal.get(COL_ASESOR_METODOLOGICO_DETALLE),
            COL_TITULO_TRABAJO: norm_str(titulo),
            COL_FECHA: pd.to_datetime(fecha),
            COL_REVISION_INICIAL: norm_str(rev_inicial),
            COL_REVISION_PLANTILLA: norm_str(rev_plantilla),
            COL_OK_REFERENCISTAS: norm_str(ok_ref),
            COL_OK_SERVICIOS: norm_str(ok_serv),
            COL_OBSERVACIONES: norm_str(obs),
            COL_ESCANEADO_TURNITIN: norm_str(esc_turnitin),
            COL_PORCENTAJE_SIMILITUD: float(similitud),
            COL_APROBACION_SIMILITUD: norm_str(aprob_sim),
            COL_PAZ_Y_SALVO: norm_str(paz_y_salvo) or "EN PROCESO",
        }

        b1, b2 = st.columns(2)

        with b1:
            if st.button("💾 Guardar registro", type="primary", key="btn_save_new"):
                if not base_row[COL_CEDULA] and not base_row[COL_NOMBRE_USUARIO]:
                    st.warning("Escribe al menos el **Nombre del usuario** o la **Cédula**.")
                else:
                    match = registro_service.find_student(base_row[COL_CEDULA], base_row[COL_NOMBRE_USUARIO])
                    if match:
                        st.warning("Este estudiante ya existe. Usa **Modificar / Adicionar asesoría**.")
                    else:
                        registro_service.create_registro(base_row, asesorias_payload)
                        st.success("Registro guardado ✔️.")
                        schedule_reset()

        with b2:
            if st.button("✏️ Modificar / Adicionar asesoría", type="secondary", key="btn_mod_add"):
                if not base_row[COL_CEDULA] and not base_row[COL_NOMBRE_USUARIO]:
                    st.warning("Escribe al menos el **Nombre del usuario** o la **Cédula** para encontrarlo.")
                else:
                    match = registro_service.find_student(base_row[COL_CEDULA], base_row[COL_NOMBRE_USUARIO])
                    if not match:
                        st.warning("No encontré al estudiante. Usa **Guardar registro** para crearlo primero.")
                    else:
                        registro_service.update_registro(match["id"], base_row, asesorias_payload)
                        st.success("Registro modificado ✔️ (asesorías adicionadas)")
                        schedule_reset()

    with colB:
        st.subheader("Vista rápida")
        df_latest = load_registro()
        st.caption("Últimos 15 registros (1 fila por estudiante)")
        show = df_latest.copy()
        if not show.empty and COL_FECHA in show.columns:
            show[COL_FECHA] = pd.to_datetime(show[COL_FECHA], errors="coerce")
        st.dataframe(show.sort_values(COL_FECHA, ascending=False).head(15), use_container_width=True)

    st.markdown("</div>", unsafe_allow_html=True)

with tab2:
    st.markdown('<div class="card">', unsafe_allow_html=True)
    st.subheader("Buscar usuario, ver total, borrar y descargar Excel")

    df_latest = load_registro()
    q = st.text_input("Buscar por nombre o cédula", placeholder="Ej: Valentina o 1032331000", key="q_search")
    if q:
        q_low = q.strip().lower()

        def match_row(row):
            n = str(row.get(COL_NOMBRE_USUARIO, "")).lower()
            c = str(row.get(COL_CEDULA, "")).lower()
            return q_low in n or q_low in c

        filtered = df_latest[df_latest.apply(match_row, axis=1)].copy()
    else:
        filtered = df_latest.copy()

    st.write(f"Registros encontrados: **{len(filtered)}**")
    cols_show = [
        col
        for col in [
            COL_NOMBRE_USUARIO,
            COL_CEDULA,
            COL_TITULO_TRABAJO,
            COL_TOTAL_ASESORIAS,
            COL_HISTORIAL_ASESORIAS,
            COL_HISTORIAL_ASESOR_RECURSOS,
            COL_PAZ_Y_SALVO,
            COL_FECHA,
        ]
        if col in filtered.columns
    ]
    if COL_FECHA in filtered.columns:
        filtered[COL_FECHA] = pd.to_datetime(filtered[COL_FECHA], errors="coerce")
    if not filtered.empty:
        st.dataframe(filtered[cols_show].sort_values(COL_FECHA, ascending=False), use_container_width=True)
    else:
        st.info("No hay registros para mostrar.")

    st.divider()
    st.markdown("### 📌 Historial del estudiante (vista detallada)")
    if len(filtered) == 0:
        st.info("Primero busca un estudiante para ver su historial.")
    else:
        options = []
        for idx, row in filtered.iterrows():
            label = f"{row.get(COL_NOMBRE_USUARIO, '')} | {row.get(COL_CEDULA, '')} | Total: {row.get(COL_TOTAL_ASESORIAS, 0)}"
            options.append((int(idx), label))
        sel = st.selectbox("Selecciona el estudiante para ver el historial", options, format_func=lambda x: x[1], key="hist_select")
        sel_idx = sel[0]
        row = df_latest.loc[sel_idx]

        cinfo1, cinfo2, cinfo3 = st.columns(3)
        with cinfo1:
            st.metric("Estudiante", str(row.get(COL_NOMBRE_USUARIO, "")))
            st.write(f"**Cédula:** {row.get(COL_CEDULA, '')}")
        with cinfo2:
            st.metric("Total asesorías", int(row.get(COL_TOTAL_ASESORIAS, 0) or 0))
            st.write(f"**Paz y salvo:** {row.get(COL_PAZ_Y_SALVO, '')}")
        with cinfo3:
            st.write("**Título trabajo de grado:**")
            st.write(str(row.get(COL_TITULO_TRABAJO, "")))

        hist_df = history_table_from_row(row)
        st.dataframe(hist_df, use_container_width=True)

        st.markdown("### 🗑️ Eliminar una asesoría (solo del historial)")
        if len(hist_df) == 0:
            st.info("Este estudiante no tiene asesorías registradas.")
        else:
            n_values = hist_df["N"].tolist()
            n_to_delete = st.selectbox("Selecciona el N de la asesoría a eliminar", n_values, key="n_to_delete")

            sel_row = hist_df[hist_df["N"] == n_to_delete].iloc[0]
            st.write("Vas a eliminar:")
            st.write(f"- **Fecha:** {sel_row.get('Fecha', '')}")
            st.write(f"- **Asesoría:** {sel_row.get('Asesoria', '')}")
            st.write(f"- **Asesor Recursos:** {sel_row.get('Asesor Recursos', '')}")
            st.write(f"- **Asesor Metodológico:** {sel_row.get('Asesor Metodologico', '')}")

            confirm_one = st.checkbox("Confirmo eliminar SOLO esta asesoría del historial", key="confirm_delete_one")
            if st.button("🗑️ Eliminar asesoría seleccionada", type="primary", disabled=not confirm_one, key="btn_delete_one"):
                success = delete_history_item(sel_idx, int(n_to_delete) - 1)
                if success:
                    st.success("Asesoría eliminada ✔️.")
                    st.rerun()
                else:
                    st.error("No fue posible eliminar la asesoría.")

        bio_h = io.BytesIO()
        with pd.ExcelWriter(bio_h, engine="xlsxwriter") as writer:
            hist_df.to_excel(writer, index=False, sheet_name="Historial")
        bio_h.seek(0)
        st.download_button(
            "📥 Descargar historial (Excel)",
            data=bio_h.read(),
            file_name=f"historial_{str(row.get(COL_CEDULA, '')).strip() or 'estudiante'}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="dl_hist_excel",
        )

    st.divider()
    st.markdown("### 🗑️ Eliminar registro (estudiante)")
    if len(filtered) == 0:
        st.info("No hay registros para eliminar con el filtro actual.")
    else:
        options_del = []
        for idx, row in filtered.iterrows():
            label = f"[{idx}] {row.get(COL_NOMBRE_USUARIO, '')} | {row.get(COL_CEDULA, '')} | Total: {row.get(COL_TOTAL_ASESORIAS, 0)}"
            options_del.append((int(idx), label))
        selected = st.selectbox("Selecciona el estudiante (fila) a eliminar", options_del, format_func=lambda x: x[1], key="del_select")
        idx_to_delete = selected[0]
        confirm = st.checkbox("Confirmo que deseo eliminar este registro", key="del_confirm")
        if st.button("🗑️ Eliminar registro seleccionado", type="primary", disabled=not confirm, key="del_btn"):
            registro_service.delete_registro(idx_to_delete)
            st.success("Registro eliminado correctamente ✔️.")
            st.rerun()

    st.divider()
    st.markdown("### 📥 Descargar Excel actual")
    st.download_button(
        label="📥 Descargar registro (Excel)",
        data=download_current_excel_bytes(),
        file_name="registro_asesorias_actual.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        key="dl_current_excel",
    )
    st.markdown("</div>", unsafe_allow_html=True)

with tab3:
    st.markdown('<div class="card">', unsafe_allow_html=True)
    st.subheader("Carga masiva (plantilla ✚ subir ✚ actualizar registros)")

    template_cols_order = load_registro().columns.tolist()
    st.markdown("**1) Descargar plantilla** (mismos campos del formulario + paz y salvo).")
    st.download_button(
        "📥 Descargar plantilla para carga masiva",
        data=bulk_template_bytes(template_cols_order),
        file_name="plantilla_carga_masiva_asesorias.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        key="dl_bulk_template",
    )

    st.markdown("**2) Subir archivo diligenciado** (crea o modifica el registro del estudiante).")
    up = st.file_uploader("Sube el Excel", type=["xlsx"], key="up_bulk")
    if up is not None:
        try:
            df_up = pd.read_excel(up)
            st.write("Vista previa (primeras 10 filas):")
            st.dataframe(df_up.head(10), use_container_width=True)
            if st.button("✅ Importar / Actualizar", type="primary", key="btn_bulk_import"):
                bulk_import(df_up)
                st.success("Importación completada ✔️.")
                st.rerun()
        except Exception as exc:
            st.error(f"No se pudo leer/importar el archivo: {exc}")
    st.markdown("</div>", unsafe_allow_html=True)
