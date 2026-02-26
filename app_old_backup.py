"""Tablero de Asesorías (Tesis) - Streamlit (v6)

✅ Mantiene listas desplegables desde el template XLSM
✅ Pestaña 1: Guardar + Modificar/Adicionar asesoría (1 fila por estudiante)
   - Limpia formulario al guardar/modificar SIN errores de session_state
   - Autocompleta por cédula (nombre/título/paz y salvo + fac/prog si coincide)
   - Soporta agregar MÚLTIPLES asesorías en un solo envío con botón ➕ (bloques dinámicos)
✅ Pestaña 2: Buscar + eliminar + descargar Excel
✅ Pestaña 3: Plantilla + carga masiva (crea o adiciona)
"""

import io
import os
from datetime import date
import pandas as pd
import streamlit as st
import openpyxl

st.set_page_config(page_title="Tablero Asesorías (Tesis)", layout="wide")

BASE_DIR = os.path.dirname(__file__)
DATA_DIR = os.path.join(BASE_DIR, "data")
TEMPLATE_PATH = os.path.join(DATA_DIR, "Control_Asesorias_Tesis_template.xlsm")
DB_PATH = os.path.join(DATA_DIR, "registro_actual.xlsx")

SHEET_REGISTRO = "Registro asesorías"
SHEET_FAC = "Data_Facultades"
SHEET_PROG = "Data_Programas"
SHEET_LISTAS = "Insumo_Listas_Desplegables"

os.makedirs(DATA_DIR, exist_ok=True)

st.markdown(
    """
    <style>
      :root{--brand:#1B8F4B;--brand-2:#146B39;--bg:#F6F7F9;--card:#FFFFFF;--muted:#6B7280;--border:#E5E7EB;}
      .stApp { background: var(--bg); }
      .muted { color: var(--muted); }
      div.stButton > button[kind="primary"]{background: var(--brand) !important;border: 1px solid var(--brand-2) !important;color: white !important;border-radius: 10px !important;padding: 0.55rem 0.9rem !important;}
      div.stButton > button[kind="secondary"]{border-radius: 10px !important;}
      .card{background: var(--card);border: 1px solid var(--border);border-radius: 14px;padding: 16px 16px 8px 16px;box-shadow: 0 1px 2px rgba(0,0,0,0.05);margin-bottom: 12px;}
      button[data-baseweb="tab"][aria-selected="true"]{ color: var(--brand) !important; }
      [data-testid="stDataFrame"]{border-radius: 14px;overflow: hidden;border: 1px solid var(--border);background: white;}
    </style>
    """,
    unsafe_allow_html=True,
)

def norm_str(x):
    if x is None:
        return None
    s = str(x).strip()
    return s if s else None

def normalize_fac_name(name: str) -> str:
    if name is None:
        return ""
    s = str(name).strip().replace(",", "")
    return "_".join([p for p in s.split() if p])

def safe_int(x, default=0):
    try:
        return int(x)
    except Exception:
        return default

def append_pipe(existing, item) -> str:
    """Agrega 'item' al historial separado por ' | ' sin duplicar. Soporta None/NaN/float."""
    item = norm_str(item)
    if not item:
        if existing is None:
            return ""
        try:
            if isinstance(existing, float) and pd.isna(existing):
                return ""
        except Exception:
            pass
        return str(existing).strip()

    try:
        if existing is None or (isinstance(existing, float) and pd.isna(existing)):
            existing = ""
        else:
            existing = str(existing).strip()
    except Exception:
        existing = "" if existing is None else str(existing).strip()

    parts = [p.strip() for p in str(existing).split(" | ") if p.strip()]
    if item not in parts:
        parts.append(item)
    return " | ".join(parts)

@st.cache_data(show_spinner=False)
def load_lists(template_path: str) -> dict:
    wb = openpyxl.load_workbook(template_path, data_only=True, keep_vba=True)
    ws = wb[SHEET_LISTAS]

    def extract_col(col_idx: int, start_row: int = 2):
        vals = []
        for r in range(start_row, ws.max_row + 1):
            v = ws.cell(r, col_idx).value
            if v is None:
                continue
            v = str(v).strip()
            if v:
                vals.append(v)
        out, seen = [], set()
        for v in vals:
            if v not in seen:
                out.append(v)
                seen.add(v)
        return out

    headers = {}
    for c in range(1, ws.max_column + 1):
        h = ws.cell(1, c).value
        if h:
            headers[str(h).strip()] = c

    lists = {name: extract_col(col_idx) for name, col_idx in headers.items()}
    df_fac = pd.read_excel(template_path, sheet_name=SHEET_FAC)
    df_prog = pd.read_excel(template_path, sheet_name=SHEET_PROG)

    fac_names = df_fac["Nombre_Facultad"].dropna().astype(str).tolist()
    fac_norm_map = {normalize_fac_name(n): n for n in fac_names}

    return {"lists": lists, "df_fac": df_fac, "df_prog": df_prog, "fac_norm_map": fac_norm_map}

EXTRA_COLS = ["Total_Asesorías", "Historial_Asesorías", "Historial_Fechas", "Historial_Asesor_Recursos", "Historial_Asesor_Metodologico", "Paz_y_Salvo"]

def ensure_db():
    if not os.path.exists(DB_PATH):
        df_template = pd.read_excel(TEMPLATE_PATH, sheet_name=SHEET_REGISTRO)
        df_empty = df_template.iloc[0:0].copy()
        for c in EXTRA_COLS:
            if c not in df_empty.columns:
                df_empty[c] = pd.Series(dtype="object")
        df_empty.to_excel(DB_PATH, index=False)

def load_registro() -> pd.DataFrame:
    ensure_db()
    df = pd.read_excel(DB_PATH)
    for c in EXTRA_COLS:
        if c not in df.columns:
            df[c] = None
    if "Fecha" in df.columns:
        df["Fecha"] = pd.to_datetime(df["Fecha"], errors="coerce")
    df["Total_Asesorías"] = pd.to_numeric(df["Total_Asesorías"], errors="coerce").fillna(0).astype(int)
    return df

def save_registro(df: pd.DataFrame):
    tmp_path = DB_PATH.replace(".xlsx", "_tmp.xlsx")
    df.to_excel(tmp_path, index=False)
    os.replace(tmp_path, DB_PATH)

def find_student_index(df: pd.DataFrame, cedula: str, nombre_usuario: str):
    cedula = norm_str(cedula)
    nombre_usuario = norm_str(nombre_usuario)

    if cedula and "Cédula" in df.columns:
        m = df[df["Cédula"].astype(str).str.strip() == cedula]
        if not m.empty:
            return int(m.index[0])

    if nombre_usuario and "Nombre_Usuario" in df.columns:
        m = df[df["Nombre_Usuario"].astype(str).str.strip().str.lower() == nombre_usuario.lower()]
        if not m.empty:
            return int(m.index[0])

    return None

def delete_registro(index_to_delete: int):
    df = load_registro()
    df = df.drop(index=index_to_delete).reset_index(drop=True)
    save_registro(df)

def download_current_excel_bytes() -> bytes:
    df = load_registro()
    bio = io.BytesIO()
    with pd.ExcelWriter(bio, engine="xlsxwriter") as writer:
        df.to_excel(writer, index=False, sheet_name="Registro")
    bio.seek(0)
    return bio.read()

def all_widget_keys():
    base = ["facultad","programa","cedula","nombre_usuario","titulo","fecha","rev_inicial","rev_plantilla","ok_ref","ok_serv",
            "esc_turnitin","similitud","aprob_sim","obs","paz_y_salvo"]
    n = int(st.session_state.get("asesorias_n", 1))
    for i in range(n):
        base += [f"asesor_rec_{i}", f"nombre_asesoria_{i}", f"modalidad_{i}", f"asesor_metodologico_{i}", f"modalidad2_{i}"]
    return base

def schedule_reset():
    st.session_state["reset_pending"] = True
    st.rerun()

def perform_reset(meta):
    for k in all_widget_keys():
        st.session_state.pop(k, None)

    st.session_state["asesorias_n"] = 1
    st.session_state["cedula"] = ""
    st.session_state["nombre_usuario"] = ""
    st.session_state["titulo"] = ""
    st.session_state["fecha"] = date.today()
    st.session_state["ok_ref"] = ""
    st.session_state["ok_serv"] = ""
    st.session_state["obs"] = ""
    st.session_state["similitud"] = 0.0
    st.session_state["paz_y_salvo"] = "En proceso"

    fac_names = meta["df_fac"]["Nombre_Facultad"].dropna().astype(str).tolist()
    st.session_state["facultad"] = fac_names[0] if fac_names else ""
    if fac_names:
        df_fac = meta["df_fac"]; df_prog = meta["df_prog"]
        fac_code = int(df_fac.loc[df_fac["Nombre_Facultad"] == fac_names[0], "Codigo_Facultad"].iloc[0])
        progs = df_prog.loc[df_prog["Código_Facultad"] == fac_code, "Nombre_Programa"].dropna().astype(str).tolist()
        st.session_state["programa"] = progs[0] if progs else ""
    else:
        st.session_state["programa"] = ""

    lists = meta["lists"]
    st.session_state["rev_inicial"] = (lists.get("Revisión Inicial") or [""])[0]
    rev_pl = lists.get("Revisión de Plantilla") or lists.get("Revisión plantilla") or [""]
    st.session_state["rev_plantilla"] = rev_pl[0] if rev_pl else ""
    st.session_state["esc_turnitin"] = (lists.get("Escaneado Turnitin") or [""])[0]
    st.session_state["aprob_sim"] = (lists.get("Aprobados PyS") or [""])[0]

    st.session_state["asesor_rec_0"] = (lists.get("Asesor_Recursos_Académicos") or [""])[0]
    st.session_state["nombre_asesoria_0"] = (lists.get("Nombre_Asesoría") or [""])[0]
    st.session_state["modalidad_0"] = (lists.get("Modalidad_Asesoría") or ["Virtual"])[0]
    st.session_state["asesor_metodologico_0"] = ""
    st.session_state["modalidad2_0"] = ""

    st.session_state["reset_pending"] = False

def ensure_dynamic_defaults(meta):
    lists = meta["lists"]
    n = int(st.session_state.get("asesorias_n", 1))
    for i in range(n):
        st.session_state.setdefault(f"asesor_rec_{i}", (lists.get("Asesor_Recursos_Académicos") or [""])[0])
        st.session_state.setdefault(f"nombre_asesoria_{i}", (lists.get("Nombre_Asesoría") or [""])[0])
        st.session_state.setdefault(f"modalidad_{i}", (lists.get("Modalidad_Asesoría") or ["Virtual"])[0])
        st.session_state.setdefault(f"asesor_metodologico_{i}", "")
        st.session_state.setdefault(f"modalidad2_{i}", "")

def add_asesoria():
    st.session_state["asesorias_n"] = int(st.session_state.get("asesorias_n", 1)) + 1

def remove_asesoria():
    n = int(st.session_state.get("asesorias_n", 1))
    if n <= 1:
        return
    last = n - 1
    for k in [f"asesor_rec_{last}", f"nombre_asesoria_{last}", f"modalidad_{last}", f"asesor_metodologico_{last}", f"modalidad2_{last}"]:
        st.session_state.pop(k, None)
    st.session_state["asesorias_n"] = n - 1

def autofill_by_cedula(meta):
    ced = st.session_state.get("cedula", "").strip()
    if not ced:
        return
    df = load_registro()
    idx = find_student_index(df, cedula=ced, nombre_usuario=None)
    if idx is None:
        return

    st.session_state["nombre_usuario"] = str(df.loc[idx, "Nombre_Usuario"] or "")
    st.session_state["titulo"] = str(df.loc[idx, "Título_Trabajo_Grado"] or "")
    st.session_state["paz_y_salvo"] = str(df.loc[idx, "Paz_y_Salvo"] or "En proceso")

    fac_norm = str(df.loc[idx, "Nombre_Facultad"] or "").strip()
    fac_display = meta["fac_norm_map"].get(fac_norm)
    if fac_display:
        st.session_state["facultad"] = fac_display
        prog = str(df.loc[idx, "Nombre_Programa"] or "").strip()
        if prog:
            st.session_state["programa"] = prog

def bulk_template_bytes(template_columns: list[str]) -> bytes:
    cols = [c for c in template_columns if c not in EXTRA_COLS]
    if "Paz_y_Salvo" not in cols:
        cols.append("Paz_y_Salvo")
    df = pd.DataFrame(columns=cols)
    bio = io.BytesIO()
    with pd.ExcelWriter(bio, engine="xlsxwriter") as writer:
        df.to_excel(writer, index=False, sheet_name="Plantilla")
    bio.seek(0)
    return bio.read()

def bulk_import(df_upload: pd.DataFrame):
    """Carga masiva: si existe estudiante => adiciona 1 asesoría; si no => crea."""
    df = load_registro()
    col8_name = "Asesor_Metodológico, Modalidad_Asesoría2, Asesor_Metodológico, Modalidad_Asesoría2"

    for _, row in df_upload.iterrows():
        ced = norm_str(row.get("Cédula"))
        nom = norm_str(row.get("Nombre_Usuario"))
        if not ced and not nom:
            continue

        fec = pd.to_datetime(row.get("Fecha"), errors="coerce")
        if pd.isna(fec):
            fec = pd.to_datetime(date.today())

        idx = find_student_index(df, cedula=ced, nombre_usuario=nom)

        # base fields (todo lo que venga en el archivo y exista en la DB)
        base_row = {}
        for c in df.columns:
            if c in EXTRA_COLS:
                continue
            if c in row.index:
                base_row[c] = row.get(c)
        if "Fecha" in base_row:
            base_row["Fecha"] = fec

        paz = norm_str(row.get("Paz_y_Salvo")) or "En proceso"
        asesor_rec = norm_str(row.get("Asesor_Recursos_Académicos"))
        nombre_asesoria = norm_str(row.get("Nombre_Asesoría"))
        campo_h = norm_str(row.get(col8_name))

        if idx is None:
            # Nuevo estudiante
            base_row["Total_Asesorías"] = 1
            base_row["Historial_Asesorías"] = nombre_asesoria or ""
            base_row["Historial_Fechas"] = fec.strftime("%Y-%m-%d")
            base_row["Historial_Asesor_Recursos"] = asesor_rec or ""
            base_row["Historial_Asesor_Metodologico"] = campo_h or ""
            base_row["Paz_y_Salvo"] = paz
            df = pd.concat([df, pd.DataFrame([base_row])], ignore_index=True)
        else:
            # Actualiza campos base (si vienen en el archivo)
            for k, v in base_row.items():
                if k in df.columns and norm_str(v) is not None:
                    df.loc[idx, k] = v

            # Adiciona 1 asesoría al historial
            df.loc[idx, "Total_Asesorías"] = safe_int(df.loc[idx, "Total_Asesorías"]) + 1
            df.loc[idx, "Historial_Asesorías"] = append_pipe(df.loc[idx, "Historial_Asesorías"], nombre_asesoria)
            df.loc[idx, "Historial_Fechas"] = append_pipe(df.loc[idx, "Historial_Fechas"], fec.strftime("%Y-%m-%d"))
            df.loc[idx, "Historial_Asesor_Recursos"] = append_pipe(df.loc[idx, "Historial_Asesor_Recursos"], asesor_rec)
            df.loc[idx, "Historial_Asesor_Metodologico"] = append_pipe(df.loc[idx, "Historial_Asesor_Metodologico"], campo_h)
            df.loc[idx, "Paz_y_Salvo"] = paz

    save_registro(df)


# ----------------------------
# Historial "bonito" (tabla por filas)
# ----------------------------
def split_hist(s):
    """Divide un historial 'a | b | c' en lista. Soporta None/NaN/float."""
    if s is None:
        return []
    try:
        if isinstance(s, float) and pd.isna(s):
            return []
    except Exception:
        pass
    s = str(s).strip()
    if not s:
        return []
    return [p.strip() for p in s.split(" | ") if p.strip()]

def history_table_from_row(row: pd.Series) -> pd.DataFrame:
    """
    Tabla bonita: 1 fila por asesoría, con columnas separadas para asesores.
    Usa:
    - Historial_Asesorías
    - Historial_Fechas
    - Historial_Asesor_Recursos
    - Historial_Asesor_Metodologico
    """
    ases = split_hist(row.get("Historial_Asesorías"))
    fechas = split_hist(row.get("Historial_Fechas"))
    rec = split_hist(row.get("Historial_Asesor_Recursos"))
    met = split_hist(row.get("Historial_Asesor_Metodologico"))

    n = max(len(ases), len(fechas), len(rec), len(met), 1)

    def get(lst, i):
        return lst[i] if i < len(lst) else ""

    data = []
    for i in range(n):
        data.append({
            "N": i + 1,
            "Fecha": get(fechas, i),
            "Asesoría": get(ases, i),
            "Asesor Recursos": get(rec, i),
            "Asesor Metodológico": get(met, i),
        })
    return pd.DataFrame(data)


def join_hist(items: list[str]) -> str:
    """Convierte lista en string 'a | b | c' limpiando vacíos."""
    items = [str(x).strip() for x in items if str(x).strip()]
    return " | ".join(items)

def delete_history_item(row: pd.Series, pos: int) -> pd.Series:
    """Elimina UNA asesoría (índice pos) de los 4 historiales alineados."""
    ases = split_hist(row.get("Historial_Asesorías"))
    fechas = split_hist(row.get("Historial_Fechas"))
    rec = split_hist(row.get("Historial_Asesor_Recursos"))
    met = split_hist(row.get("Historial_Asesor_Metodologico"))

    max_len = max(len(ases), len(fechas), len(rec), len(met))
    if pos < 0 or pos >= max_len:
        return row

    if pos < len(ases): ases.pop(pos)
    if pos < len(fechas): fechas.pop(pos)
    if pos < len(rec): rec.pop(pos)
    if pos < len(met): met.pop(pos)

    row["Historial_Asesorías"] = join_hist(ases)
    row["Historial_Fechas"] = join_hist(fechas)
    row["Historial_Asesor_Recursos"] = join_hist(rec)
    row["Historial_Asesor_Metodologico"] = join_hist(met)

    total = safe_int(row.get("Total_Asesorías", 0))
    row["Total_Asesorías"] = max(total - 1, 0)

    # Actualiza campos principales con el primer elemento restante (si hay)
    if row["Total_Asesorías"] <= 0:
        row["Nombre_Asesoría"] = ""
        row["Asesor_Recursos_Académicos"] = ""
        row["Fecha"] = pd.NaT
    else:
        if len(ases) > 0:
            row["Nombre_Asesoría"] = ases[0]
        if len(rec) > 0:
            row["Asesor_Recursos_Académicos"] = rec[0]
        if len(fechas) > 0:
            row["Fecha"] = pd.to_datetime(fechas[0], errors="coerce")

    return row

# ---------- UI ----------
if not os.path.exists(TEMPLATE_PATH):
    st.error(f"No encuentro el template en: {TEMPLATE_PATH}\\n\\n✅ Debe existir: data/Control_Asesorias_Tesis_template.xlsm")
    st.stop()

meta = load_lists(TEMPLATE_PATH)
lists = meta["lists"]
df_fac = meta["df_fac"]
df_prog = meta["df_prog"]

st.title("Asesorías – Trabajo de grado")


tab1, tab2, tab3 = st.tabs(["➕ Registrar asesoría", "🔎 Consultar usuario", "⬆️ Carga masiva"])

with tab1:

    if st.session_state.get("reset_pending", False) or "init_done" not in st.session_state:
        perform_reset(meta)
        st.session_state["init_done"] = True

    ensure_dynamic_defaults(meta)

    colA, colB = st.columns([1.15, 0.85], gap="large")

    with colA:
        st.subheader("Formulario de registro")

        fac_names = df_fac["Nombre_Facultad"].dropna().astype(str).tolist()
        fac_display = st.selectbox("Facultad", fac_names, index=0, key="facultad")

        fac_code = int(df_fac.loc[df_fac["Nombre_Facultad"] == fac_display, "Codigo_Facultad"].iloc[0])
        progs = df_prog.loc[df_prog["Código_Facultad"] == fac_code, "Nombre_Programa"].dropna().astype(str).tolist()
        if not progs: progs = [""]
        prog_display = st.selectbox("Programa", progs, index=0, key="programa")

        c1, c2 = st.columns(2)
        with c1:
            cedula = st.text_input("Cédula", placeholder="Ej: 1032331000", key="cedula",
                                   on_change=lambda: autofill_by_cedula(meta))
        with c2:
            nombre_usuario = st.text_input("Nombre del usuario", placeholder="Ej: Juan Pérez", key="nombre_usuario")

        titulo = st.text_input("Título trabajo de grado", key="titulo")
        fecha = st.date_input("Fecha", key="fecha")

        st.markdown("### Asesorías a registrar")
        n = int(st.session_state.get("asesorias_n", 1))

        cadd, crem, _ = st.columns([0.25, 0.25, 0.5])
        with cadd:
            st.button("➕ Agregar asesoría", on_click=add_asesoria, key="btn_add_asesoria")
        with crem:
            st.button("➖ Quitar última", on_click=remove_asesoria, disabled=(n <= 1), key="btn_remove_asesoria")

        col8_name = "Asesor_Metodológico, Modalidad_Asesoría2, Asesor_Metodológico, Modalidad_Asesoría2"
        asesorias_payload = []

        for i in range(n):
            with st.expander(f"Asesoría #{i+1}", expanded=(i == 0)):
                asesor_rec_i = st.selectbox("Asesor Recursos Académicos", lists.get("Asesor_Recursos_Académicos", [""]),
                                            key=f"asesor_rec_{i}")
                nombre_asesoria_i = st.selectbox("Nombre de la asesoría", lists.get("Nombre_Asesoría", [""]),
                                                 key=f"nombre_asesoria_{i}")
                modalidad_i = st.selectbox("Modalidad", lists.get("Modalidad_Asesoría", ["Virtual","Presencial"]),
                                           key=f"modalidad_{i}")

                st.markdown("**Asesor metodológico**")
                asesor_met_i = st.text_input("Asesor metodológico (digita el nombre)", placeholder="Ej: Carlos Rodríguez",
                                             key=f"asesor_metodologico_{i}")
                modalidad2_i = st.selectbox("Modalidad asesoría ", ["", "Virtual", "Presencial"],
                                            key=f"modalidad2_{i}")

                if asesor_met_i and modalidad2_i:
                    campo_h_i = f"{asesor_met_i.strip()}, {modalidad2_i}"
                elif asesor_met_i:
                    campo_h_i = asesor_met_i.strip()
                else:
                    campo_h_i = None

                asesorias_payload.append({
                    "Asesor_Recursos_Académicos": norm_str(asesor_rec_i),
                    "Nombre_Asesoría": norm_str(nombre_asesoria_i),
                    "Modalidad del Programa": norm_str(modalidad_i),
                    col8_name: norm_str(campo_h_i),
                })

        c3, c4 = st.columns(2)
        with c3:
            rev_inicial = st.selectbox("Revisión inicial", lists.get("Revisión Inicial", [""]), key="rev_inicial")
            rev_pl_opts = lists.get("Revisión de Plantilla") or lists.get("Revisión plantilla") or [""]
            rev_plantilla = st.selectbox("Revisión plantilla", rev_pl_opts, key="rev_plantilla")
            ok_ref = st.text_input("Ok_Referencistas", placeholder="Ej: listo / pendiente", key="ok_ref")
        with c4:
            ok_serv = st.text_input("OK_Servicios", placeholder="Ej: listo / pendiente", key="ok_serv")
            esc_turnitin = st.selectbox("Escaneado Turnitin", lists.get("Escaneado Turnitin", [""]), key="esc_turnitin")
            similitud = st.number_input("% similitud", min_value=0.0, max_value=100.0, step=0.1, key="similitud")

        aprob_sim = st.selectbox("Aprobación similitud", lists.get("Aprobados PyS", [""]), key="aprob_sim")
        obs = st.text_area("Observaciones", height=120, key="obs")
        paz_y_salvo = st.selectbox("Estudiante apto para paz y salvo", ["En proceso", "Si", "No"], index=0, key="paz_y_salvo")

        # principal
        principal = asesorias_payload[0] if asesorias_payload else {}

        fac_norm = normalize_fac_name(fac_display)
        base_row = {
            "Nombre_Facultad": fac_norm,
            "Nombre_Programa": prog_display,
            "Cédula": norm_str(cedula),
            "Nombre_Usuario": norm_str(nombre_usuario),
            "Asesor_Recursos_Académicos": principal.get("Asesor_Recursos_Académicos"),
            "Nombre_Asesoría": principal.get("Nombre_Asesoría"),
            "Modalidad del Programa": principal.get("Modalidad del Programa"),
            col8_name: principal.get(col8_name),
            "Título_Trabajo_Grado": norm_str(titulo),
            "Fecha": pd.to_datetime(fecha),
            "Revisión Inicial": norm_str(rev_inicial),
            "Revisión plantilla": norm_str(rev_plantilla),
            "Ok_Referencistas": norm_str(ok_ref),
            "OK_Servicios": norm_str(ok_serv),
            "Observaciones": norm_str(obs),
            "Escaneado Turnitin": norm_str(esc_turnitin),
            "% similitud": float(similitud),
            "Aprobación_Similitud": norm_str(aprob_sim),
            "Paz_y_Salvo": norm_str(paz_y_salvo) or "En proceso",
        }

        b1, b2 = st.columns(2)

        with b1:
            if st.button("✅ Guardar registro", type="primary", key="btn_save_new"):
                if not base_row["Cédula"] and not base_row["Nombre_Usuario"]:
                    st.warning("Escribe al menos el **Nombre del usuario** o la **Cédula**.")
                else:
                    df_current = load_registro()
                    idx = find_student_index(df_current, base_row["Cédula"], base_row["Nombre_Usuario"])
                    if idx is not None:
                        st.warning("Este estudiante ya existe. Usa **Modificar / Adicionar asesoría**.")
                    else:
                        total_new = len(asesorias_payload) if asesorias_payload else 0
                        base_row["Total_Asesorías"] = total_new

                        hist_asesorias = ""
                        hist_rec = ""
                        hist_met = ""
                        hist_fechas = ""
                        for a in asesorias_payload:
                            hist_asesorias = append_pipe(hist_asesorias, a.get("Nombre_Asesoría"))
                            hist_fechas = append_pipe(hist_fechas, pd.to_datetime(fecha).strftime("%Y-%m-%d"))
                            hist_rec = append_pipe(hist_rec, a.get("Asesor_Recursos_Académicos"))
                            hist_met = append_pipe(hist_met, a.get(col8_name))

                        base_row["Historial_Asesorías"] = hist_asesorias
                        base_row["Historial_Asesor_Recursos"] = hist_rec
                        base_row["Historial_Asesor_Metodologico"] = hist_met
                        base_row["Historial_Fechas"] = hist_fechas

                        df2 = pd.concat([df_current, pd.DataFrame([base_row])], ignore_index=True)
                        save_registro(df2)

                        st.success("Registro guardado ✅")
                        schedule_reset()

        with b2:
            if st.button("📝 Modificar / Adicionar asesoría", type="secondary", key="btn_mod_add"):
                if not base_row["Cédula"] and not base_row["Nombre_Usuario"]:
                    st.warning("Escribe al menos el **Nombre del usuario** o la **Cédula** para encontrarlo.")
                else:
                    df_current = load_registro()
                    idx = find_student_index(df_current, base_row["Cédula"], base_row["Nombre_Usuario"])
                    if idx is None:
                        st.warning("No encontré al estudiante. Usa **Guardar registro** para crearlo primero.")
                    else:
                        for k, v in base_row.items():
                            if k in df_current.columns and norm_str(v) is not None:
                                df_current.loc[idx, k] = v

                        add_count = len(asesorias_payload) if asesorias_payload else 0
                        df_current.loc[idx, "Total_Asesorías"] = safe_int(df_current.loc[idx, "Total_Asesorías"]) + add_count

                        for a in asesorias_payload:
                            df_current.loc[idx, "Historial_Asesorías"] = append_pipe(df_current.loc[idx, "Historial_Asesorías"], a.get("Nombre_Asesoría"))
                            df_current.loc[idx, "Historial_Asesor_Recursos"] = append_pipe(df_current.loc[idx, "Historial_Asesor_Recursos"], a.get("Asesor_Recursos_Académicos"))
                            df_current.loc[idx, "Historial_Asesor_Metodologico"] = append_pipe(df_current.loc[idx, "Historial_Asesor_Metodologico"], a.get(col8_name))
                            df_current.loc[idx, "Historial_Fechas"] = append_pipe(df_current.loc[idx, "Historial_Fechas"], pd.to_datetime(fecha).strftime("%Y-%m-%d"))

                        df_current.loc[idx, "Paz_y_Salvo"] = base_row["Paz_y_Salvo"]
                        save_registro(df_current)

                        st.success("Registro modificado ✅ (asesorías adicionadas)")
                        schedule_reset()

    with colB:
        st.subheader("Vista rápida")
        df_latest = load_registro()
        st.caption("Últimos 15 registros (1 fila por estudiante)")
        show = df_latest.copy()
        if "Fecha" in show.columns:
            show["Fecha"] = pd.to_datetime(show["Fecha"], errors="coerce")
        st.dataframe(show.sort_values("Fecha", ascending=False).tail(15), use_container_width=True)


with tab2:
    st.subheader("Buscar usuario, ver total, borrar y descargar Excel")

    df_latest = load_registro()
    q = st.text_input("Buscar por nombre o cédula", placeholder="Ej: Valentina o 1032331000", key="q_search")
    if q:
        q_low = q.strip().lower()
        def match_row(row):
            n = str(row.get("Nombre_Usuario", "")).lower()
            c = str(row.get("Cédula", "")).lower()
            return q_low in n or q_low in c
        filtered = df_latest[df_latest.apply(match_row, axis=1)].copy()
    else:
        filtered = df_latest.copy()

    st.write(f"Registros encontrados: **{len(filtered)}**")
    cols_show = [c for c in ["Nombre_Usuario","Cédula","Título_Trabajo_Grado","Total_Asesorías","Historial_Asesorías","Historial_Asesor_Recursos","Paz_y_Salvo","Fecha"] if c in filtered.columns]
    if "Fecha" in filtered.columns:
        filtered["Fecha"] = pd.to_datetime(filtered["Fecha"], errors="coerce")
    st.dataframe(filtered[cols_show].sort_values("Fecha", ascending=False), use_container_width=True)

    st.divider()
    st.markdown("### 📌 Historial del estudiante (vista detallada)")
    if len(filtered) == 0:
        st.info("Primero busca un estudiante para ver su historial.")
    else:
        # Selector del estudiante dentro del resultado filtrado
        opts = []
        for idx, row in filtered.iterrows():
            opts.append((int(idx), f"{row.get('Nombre_Usuario','')} | {row.get('Cédula','')} | Total: {row.get('Total_Asesorías',0)}"))
        sel = st.selectbox("Selecciona el estudiante para ver el historial", opts, format_func=lambda x: x[1], key="hist_select")
        sel_idx = sel[0]
        row = df_latest.loc[sel_idx]

        cinfo1, cinfo2, cinfo3 = st.columns(3)
        with cinfo1:
            st.metric("Estudiante", str(row.get("Nombre_Usuario","")))
            st.write(f"**Cédula:** {row.get('Cédula','')}")
        with cinfo2:
            st.metric("Total asesorías", int(row.get("Total_Asesorías", 0) or 0))
            st.write(f"**Paz y salvo:** {row.get('Paz_y_Salvo','')}")
        with cinfo3:
            st.write("**Título trabajo de grado:**")
            st.write(str(row.get("Título_Trabajo_Grado","")))

        hist_df = history_table_from_row(row)
        st.dataframe(hist_df, use_container_width=True)

        st.markdown("### 🧹 Eliminar una asesoría (solo del historial)")
        if len(hist_df) == 0:
            st.info("Este estudiante no tiene asesorías registradas.")
        else:
            n_values = hist_df["N"].tolist()
            n_to_delete = st.selectbox("Selecciona el N de la asesoría a eliminar", n_values, key="n_to_delete")

            sel_row = hist_df[hist_df["N"] == n_to_delete].iloc[0]
            st.write("Vas a eliminar:")
            st.write(f"- **Fecha:** {sel_row.get('Fecha','')}")
            st.write(f"- **Asesoría:** {sel_row.get('Asesoría','')}")
            st.write(f"- **Asesor Recursos:** {sel_row.get('Asesor Recursos','')}")
            st.write(f"- **Asesor Metodológico:** {sel_row.get('Asesor Metodológico','')}")

            confirm_one = st.checkbox("Confirmo eliminar SOLO esta asesoría del historial", key="confirm_delete_one")
            if st.button("🗑 Eliminar asesoría seleccionada", type="primary", disabled=not confirm_one, key="btn_delete_one"):
                df_all = load_registro()
                pos = int(n_to_delete) - 1
                updated = delete_history_item(df_all.loc[sel_idx].copy(), pos)
                df_all.loc[sel_idx] = updated
                save_registro(df_all)
                st.success("Asesoría eliminada ✅")
                st.rerun()


        # Descarga del historial del estudiante
        bio_h = io.BytesIO()
        with pd.ExcelWriter(bio_h, engine="xlsxwriter") as writer:
            hist_df.to_excel(writer, index=False, sheet_name="Historial")
        bio_h.seek(0)
        st.download_button(
            "⬇️ Descargar historial (Excel)",
            data=bio_h.read(),
            file_name=f"historial_{str(row.get('Cédula','')).strip() or 'estudiante'}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="dl_hist_excel"
        )

    st.divider()
    st.markdown("### 🗑 Eliminar registro (estudiante)")
    if len(filtered) == 0:
        st.info("No hay registros para eliminar con el filtro actual.")
    else:
        options = []
        for idx, row in filtered.iterrows():
            options.append((int(idx), f"[{idx}] {row.get('Nombre_Usuario','')} | {row.get('Cédula','')} | Total: {row.get('Total_Asesorías',0)}"))
        selected = st.selectbox("Selecciona el estudiante (fila) a eliminar", options, format_func=lambda x: x[1], key="del_select")
        idx_to_delete = selected[0]
        confirm = st.checkbox("Confirmo que deseo eliminar este registro", key="del_confirm")
        if st.button("❌ Eliminar registro seleccionado", type="primary", disabled=not confirm, key="del_btn"):
            delete_registro(idx_to_delete)
            st.success("Registro eliminado correctamente ✅")
            st.rerun()

    st.divider()
    st.markdown("### ⬇️ Descargar Excel actual")
    st.download_button(
        label="⬇️ Descargar registro (Excel)",
        data=download_current_excel_bytes(),
        file_name="registro_asesorias_actual.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        key="dl_current_excel",
    )

with tab3:
    st.subheader("Carga masiva (plantilla → subir → actualizar registros)")

    template_cols_order = load_registro().columns.tolist()
    st.markdown("**1) Descargar plantilla** (mismos campos del formulario + paz y salvo).")
    st.download_button(
        "⬇️ Descargar plantilla para carga masiva",
        data=bulk_template_bytes(template_cols_order),
        file_name="plantilla_carga_masiva_asesorias.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        key="dl_bulk_template",
    )

    st.markdown("**2) Subir archivo diligenciado** (crea o modifica el registro del estudiante).")
    up = st.file_uploader("Sube el Excel", type=["xlsx"], key="up_bulk")
    if up is not None:
        try:
            df_up = pd.read_excel(up)
            st.write("Vista previa (primeras 10 filas):")
            st.dataframe(df_up.head(10), use_container_width=True)
            if st.button("🚀 Importar / Actualizar", type="primary", key="btn_bulk_import"):
                bulk_import(df_up)
                st.success("Importación completada ✅")
                st.rerun()
        except Exception as e:
            st.error(f"No se pudo leer/importar el archivo: {e}")
